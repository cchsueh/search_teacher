import requests
import time
from openpyxl import Workbook
from bs4 import BeautifulSoup

SCHOOL_URL = 'https://www.nccu.edu.tw/'


# 爬取師資資料
def search_teacher(query):
    url = 'https://' + query + '.nccu.edu.tw/members/teacher.php'
    headers = {'User-Agent': 'mozilla/5.0 (Linux; Android 6.0.1; '
                             'Nexus 5x build/mtc19t applewebkit/537.36 (KHTML, like Gecko) '
                             'Chrome/51.0.2702.81 Mobile Safari/537.36'}
    resp = requests.get(url, headers=headers)
    resp.encoding = 'utf-8'
    soup = BeautifulSoup(resp.text, 'html5lib')
    school, department = soup.title.text.split()
    info_data = []

    info_div = soup.find_all('div', 'info')
    for info in info_div:
        mails = info.find_all('li', 'mail')
        if mails:
            for mail in mails:
               mail = mail.find('a')['href'].replace('mailto:', '')
            # 職稱
            job_title = info.find_all('li')[0].text.replace('職稱:', '')
            # 姓名
            name = info.find_all('li')[1].text.replace('姓名:', '')

            data = {
                'school': school,
                'department': department,
                'job_title': job_title,
                'name': name,
                'mail': mail,
            }
        else:
            continue
        info_data.append(data)
    return info_data


if __name__ == '__main__':
    # 搜尋科系
    query = 'chinese'
    today = time.strftime('%m-%d')
    info_data = search_teacher(query)
    # 建立Excel檔
    wb = Workbook()
    ws = wb.active
    ws.title = '中文系'
    ws['A1'] = '學校'
    ws['B1'] = '科系'
    ws['C1'] = '職稱'
    ws['D1'] = '姓名'
    ws['E1'] = 'E-MAIL'

    for d in info_data:
        ws.append([d['school'], d['department'], d['job_title'], d['name'], d['mail']])

    wb.save(today + '_nccu_' + query + '_teacher_list.xlsx')